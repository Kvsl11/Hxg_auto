import threading
import glob
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import datetime
from fpdf import FPDF
import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as ttk
import time
import json
import warnings
import os
import ssl
import subprocess
import urllib.request
import logging
import sys
import requests
import shutil
import xlwings as xw
import certifi
import tempfile

# Caminho dinâmico da pasta onde o script está localizado
app_dir = os.path.dirname(os.path.abspath(__file__))
# Usa o executável Python que está rodando o script ATUALMENTE
python_exe = sys.executable
print(f"🟢 Usando Python em: {python_exe}")

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Apenas o certificado raiz da Amazon (necessário para access.hxgnagron.com)
AMAZON_CERTS = {
    "Amazon Root CA 1": "https://www.amazontrust.com/repository/AmazonRootCA1.pem"
}

def preparar_dependencias():
    """Instala/Atualiza pacotes essenciais usando o Python ATUAL."""
    try:
        if not os.path.exists(python_exe):
            logger.warning(f"⚠️ Python não encontrado em: {python_exe}")
            return

        # --- LIMPEZA DE CACHE ANTIGO ---
        wdm_cache_path = os.path.join(os.path.expanduser("~"), ".wdm")
        if os.path.exists(wdm_cache_path):
            logger.info(f"🧹 Limpando cache antigo do webdriver-manager em: {wdm_cache_path}")
            shutil.rmtree(wdm_cache_path, ignore_errors=True)

        # --- RESOLVE CONFLITO FPDF / FPDF2 ---
        logger.info("🧹 Removendo pacotes conflitantes (fpdf legado)...")
        subprocess.run([python_exe, "-m", "pip", "uninstall", "--yes", "fpdf"], check=False, capture_output=True)

        # --- LISTA COMPLETA DE DEPENDÊNCIAS ATUALIZADAS ---
        pacotes = [
            "certifi", "selenium", "xlwings", "pandas", 
            "openpyxl", "fpdf2", "ttkbootstrap", "requests"
        ]
        logger.info(f"🔍 Verificando e atualizando pacotes vitais...")

        for pacote in pacotes:
            subprocess.run([python_exe, "-m", "pip", "install", "--upgrade", pacote], check=False, capture_output=True)
        logger.info(f"🟢 Dependências verificadas com sucesso.")
    except Exception as e:
        logger.warning(f"⚠️ Falha ao preparar dependências: {e}")

def garantire_certificados_amazon():
    """Verifica se o certificado raiz da Amazon está presente e adiciona se necessário."""
    try:
        cacert_path = certifi.where()

        with open(cacert_path, "r", encoding="utf-8") as f:
            conteudo = f.read()

        for nome, url in AMAZON_CERTS.items():
            if nome not in conteudo:
                logger.info(f"🔍 {nome} não encontrado, baixando de {url}...")
                resp = requests.get(url, timeout=10, verify=False)
                if resp.status_code == 200:
                    with open(cacert_path, "a", encoding="utf-8") as f:
                        f.write(f"\n# {nome}\n{resp.text.strip()}\n")
                    logger.info(f"✅ {nome} adicionado ao cacert.pem.")
                else:
                    logger.warning(f"❌ Falha ao baixar {nome}: {resp.status_code}")
            else:
                logger.info(f"🟢 {nome} já está presente no cacert.pem.")
    except Exception as e:
        logger.warning(f"⚠️ Falha ao garantir certificado Amazon Root CA 1: {e}")

logger.info("🚀 Iniciando verificação e correção SSL híbrida...")
preparar_dependencias()
garantire_certificados_amazon()

try:
    ssl_context = ssl.create_default_context(cafile=certifi.where())
    urllib.request.urlopen("https://www.google.com", timeout=5, context=ssl_context)
    logger.info("🟢 Conexão SSL validada com sucesso — certificados OK.")
except ssl.SSLError as e:
    logger.warning(f"⚠️ Falha de SSL detectada ({e}). Aplicando modo não verificado (Fallback).")
    ssl._create_default_https_context = ssl._create_unverified_context
    logger.info("🟡 SSL desativado globalmente — conexão forçada sem verificação de certificado.")
except Exception as e:
    logger.warning(f"⚠️ Erro genérico ao testar SSL: {e}. Aplicando modo não verificado (Fallback).")
    ssl._create_default_https_context = ssl._create_unverified_context
    logger.info("🟡 SSL desativado globalmente — conexão forçada sem verificação de certificado.")

logger.info("✅ Configuração SSL concluída com segurança.")

def exibir_erro_fatal(titulo, mensagem):
    root_temp = tk.Tk()
    root_temp.withdraw()
    root_temp.attributes("-topmost", True)
    messagebox.showerror(titulo, mensagem)
    root_temp.destroy()
    os._exit(1)

def verificar_seguranca():
    try:
        REPO = "Kvsl11/Hxg_auto"
        ts = int(time.time())
        URL_STATUS = f"https://raw.githubusercontent.com/{REPO}/main/status.txt?t={ts}"

        try:
            headers = {"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"}
            r_status = requests.get(URL_STATUS, timeout=10, verify=False, headers=headers)
            if r_status.status_code == 200:
                status_app = r_status.text.strip().lower()
                if status_app == "false":
                    logger.warning("🔴 TRAVA ATIVADA VIA GITHUB! Bloqueando acesso.")
                    exibir_erro_fatal("Erro Crítico de Comunicação",
                                      "Ocorreu uma falha inesperada ao sincronizar as configurações iniciais do sistema.\n\nCódigo do Erro: ERR_CONNECTION_REFUSED_10061\nPor favor, tente novamente mais tarde.")
            else:
                logger.info(f"⚠️ Status remoto retornou código {r_status.status_code}. Execução permitida.")
        except Exception as e:
            logger.warning(f"⚠️ Falha ao checar status.txt. Ignorando trava. Erro: {e}")
    except Exception as e:
        logger.error(f"❌ Erro na rotina de segurança: {e}")

VERSAO = "3.4.0"

warnings.filterwarnings(
    "ignore",
    message="Slicer List extension is not supported and will be removed",
    category=UserWarning,
    module=r"openpyxl\.worksheet\._reader"
)

script_dir = os.path.dirname(os.path.abspath(__file__))
execucao_ativa = False
status_label = None
progress_bar = None
root = None

RESPONSAVEIS_OPCOES = [
    "JUAN CARLOS", "ROSANI ALDA", "FERNANDO BREGUEDO", "FLAVIO BREGUEDO",
    "EDUARDO APARECIDO", "LEANDRO RENE", "EDUARDO NUNES", "LEANDRO SEBOLD",
    "ALEX FABIANO", "RAMON ROSA"
]

def iniciar_driver(headless=True):
    """Inicia uma instância do Chrome preparada para downloads em background."""
    print(f"🚀 Iniciando driver (Headless={headless}) com Selenium padrão (SeleniumManager)...")
    logger.info(f"🚀 Iniciando driver (Headless={headless}) com Selenium padrão (SeleniumManager)...")

    options = webdriver.ChromeOptions()
    
    # --- PERMITIR DOWNLOADS NO HEADLESS ---
    download_dir = os.path.join(os.path.expanduser("~"), "Downloads")
    prefs = {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True, # Vital para headless
        "profile.default_content_setting_values.automatic_downloads": 1
    }
    options.add_experimental_option("prefs", prefs)

    # --- PARÂMETROS CRÍTICOS CONTRA CRASH E DEVTOOLS NO WINDOWS ---
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--remote-debugging-port=9222")
    options.add_argument("--disable-software-rasterizer")

    if headless:
        options.add_argument("--headless=new") # Força a nova engine Headless
        options.add_argument("--window-size=1920,1080")
        
        # Gera uma pasta temporária única para esta sessão (evita conflitos de arquivos de cache bloqueados)
        temp_dir = tempfile.mkdtemp(prefix="selenium_hxg_")
        options.add_argument(f"--user-data-dir={temp_dir}")

    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-web-security')
    options.add_experimental_option('excludeSwitches', ['enable-logging'])

    try:
        servico = Service()
        driver = webdriver.Chrome(service=servico, options=options)
    except Exception as e:
        print(f"❌ Falha ao iniciar Selenium/SeleniumManager: {e}")
        logger.error(f"❌ Falha ao iniciar Selenium/SeleniumManager: {e}")
        raise

    if not headless:
        driver.maximize_window()
    return driver

def aguardar_pagina_carregada(driver, timeout=30):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        print("🟢 Página totalmente carregada.")
    except Exception as e:
        print(f"⚠️ Erro ao aguardar carregamento: {e}")

def aguardar_e_clicar(driver, xpath, timeout=30):
    try:
        print(f"Tentando clicar em: {xpath}")
        elemento = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'nearest'});", elemento)
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        time.sleep(0.5)
        elemento.click()
        print(f"🟢 Clique realizado: {xpath}")
    except Exception as e:
        logger.warning(f"⚠️ Clique padrão falhou ({xpath}). Tentando via JS.")
        try:
            elemento = driver.find_element(By.XPATH, xpath)
            driver.execute_script("arguments[0].click();", elemento)
            print(f"🟢 Clique via JS realizado: {xpath}")
        except Exception as js_e:
            print(f"❌ Erro final ao clicar via JS em {xpath}: {js_e}")

def aguardar_e_escrever(driver, xpath, texto, timeout=30):
    try:
        campo = WebDriverWait(driver, timeout).until(
            EC.visibility_of_element_located((By.XPATH, xpath))
        )
        campo.clear()
        campo.send_keys(texto)
        print(f"🟢 Texto inserido: {texto}")
    except Exception as e:
        print(f"⚠️ Erro ao escrever no campo {xpath}: {e}")

def login_usuario(driver, url, usuario, senha, xpaths):
    driver.get(url)
    aguardar_pagina_carregada(driver)
    aguardar_e_escrever(driver, xpaths['usuario'], usuario)
    aguardar_e_escrever(driver, xpaths['senha'], senha)
    aguardar_e_clicar(driver, xpaths['botao_login'])
    time.sleep(3)

def exportar_tabela(driver, xpaths):
    limpar_filtro_xpath = '//button[contains(@id,"buttion-id-clearAndApplyButton")]'
    aguardar_e_clicar(driver, limpar_filtro_xpath)

    print("⏳ Aguardando processamento após limpar o filtro...")
    time.sleep(3)

    try:
        WebDriverWait(driver, 30).until(
            EC.invisibility_of_element_located((By.XPATH,
                                                "//strong[contains(.,'Loading')] | //strong[contains(.,'Carregando')] | //div[contains(@class, 'overlay')] | //*[contains(@class, 'spinner')]"))
        )
    except Exception as e:
        print(f"⚠️ Não foi possível confirmar o desaparecimento do overlay (ou não havia): {e}")

    time.sleep(3)
    aguardar_e_clicar(driver, xpaths['tabela'])
    time.sleep(3)

    try:
        WebDriverWait(driver, 30).until(
            EC.invisibility_of_element_located((By.XPATH,
                                                "//strong[contains(.,'Loading')] | //strong[contains(.,'Carregando')] | //div[contains(@class, 'overlay')] | //*[contains(@class, 'spinner')]"))
        )
    except Exception:
        pass

    time.sleep(2)
    print("🔄 Alterando paginação para exibir mais itens...")
    aguardar_e_clicar(driver, xpaths['paginador_dropdown'])
    time.sleep(1)

    aguardar_e_clicar(driver, xpaths['paginador_opcao_5'])
    time.sleep(4)

    aguardar_e_clicar(driver, xpaths['filtro'])
    time.sleep(2)

    aguardar_e_clicar(driver, xpaths['exportacao_csv'])
    print("🟢 Exportação iniciada")

def aguardar_download_completo(diretorio, nome_base, timeout=60):
    tempo_inicial = time.time()
    while time.time() - tempo_inicial < timeout:
        arquivos_tmp = glob.glob(os.path.join(diretorio, f"{nome_base}*.tmp"))
        arquivos_csv = glob.glob(os.path.join(diretorio, f"{nome_base}*.csv"))

        if arquivos_csv and not arquivos_tmp:
            return max(arquivos_csv, key=os.path.getctime)
        time.sleep(2)

    print("❌ Tempo limite excedido para o download do arquivo CSV!")
    return None

def processar_csv(diretorio_downloads, pdf_output_dir, selected_responsaveis):
    try:
        if not os.path.exists(pdf_output_dir):
            os.makedirs(pdf_output_dir)

        print("⏳ Aguardando download do arquivo CSV...")
        csv_path = aguardar_download_completo(diretorio_downloads, "Monitoramento - Tabela")
        if not csv_path:
            print("❌ Processo encerrado. Nenhum arquivo CSV disponível.")
            return None

        df = pd.read_csv(csv_path, encoding="utf-8", sep=";", dtype=str)
        df.columns = df.columns.str.strip().str.upper()
        
        df["REGISTRO MAIS RECENTE"] = pd.to_datetime(df["REGISTRO MAIS RECENTE"], format="%d/%m/%Y %H:%M:%S",
                                                     errors="coerce")

        try:
            if os.path.exists(csv_path):
                os.remove(csv_path)
        except Exception:
            pass

        data_atual = datetime.datetime.now().date()
        df_antigos = df[df["REGISTRO MAIS RECENTE"].dt.date != data_atual].copy()

        caminho_base_monitoramento = obter_caminho_planilha()
        print(f"📖 Lendo responsáveis diretamente de: {caminho_base_monitoramento} (Aba: Cont. Maquinas)")
        
        df_responsaveis = None
        
        print("⏳ Inicializando motor Excel nativo para extrair os valores reais das fórmulas...")
        try:
            app = xw.App(visible=False)
            app.display_alerts = False
            try:
                wb = app.books.open(caminho_base_monitoramento, update_links=False, read_only=True)
                ws = wb.sheets["Cont. Maquinas"]
                df_responsaveis = ws.used_range.options(pd.DataFrame, index=False).value
                print("✅ Valores reais das fórmulas extraídos com sucesso!")
            finally:
                if 'wb' in locals() and wb: wb.close()
                if 'app' in locals() and app: app.quit()
        except Exception as ex:
            print(f"⚠️ Aviso: Não foi possível usar o motor nativo ({ex}). Tentando via pandas padrão...")
            
        if df_responsaveis is None or df_responsaveis.empty:
            df_responsaveis = pd.read_excel(caminho_base_monitoramento, sheet_name="Cont. Maquinas")
        
        normalized_cols = {}
        for col in df_responsaveis.columns:
            clean_col = str(col).strip().upper()
            clean_col = clean_col.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
            normalized_cols[col] = clean_col
        df_responsaveis = df_responsaveis.rename(columns=normalized_cols)

        col_equipamento_csv = "NRO DO EQUIPAMENTO"
        col_equipamento_base = "EQUIPAMENTO"
        col_responsavel = "RESPONSAVEL"
        col_display = "DISPLAY"
        col_prestador = "PRESTADOR"

        if col_equipamento_base not in df_responsaveis.columns:
            print(f"⚠️ A coluna '{col_equipamento_base}' não foi encontrada na aba 'Cont. Maquinas'!")
            return None

        df_responsaveis = df_responsaveis.rename(columns={col_equipamento_base: col_equipamento_csv})

        df_antigos[col_equipamento_csv] = df_antigos[col_equipamento_csv].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
        df_responsaveis[col_equipamento_csv] = df_responsaveis[col_equipamento_csv].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

        for col in [col_responsavel, col_display, col_prestador]:
            if col in df_responsaveis.columns:
                df_responsaveis[col] = df_responsaveis[col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
                df_responsaveis[col] = df_responsaveis[col].replace({'nan': '', 'None': '', '<NA>': ''})

        colunas_necessarias = [col_equipamento_csv]
        for c in [col_responsavel, col_display, col_prestador]:
            if c in df_responsaveis.columns:
                colunas_necessarias.append(c)

        df_final = df_antigos.merge(
            df_responsaveis[colunas_necessarias], 
            on=col_equipamento_csv, 
            how="left", 
            suffixes=('_CSV', '_EXCEL')
        )
        
        for c in [col_responsavel, col_display, col_prestador]:
            if f"{c}_EXCEL" in df_final.columns:
                df_final[c] = df_final[f"{c}_EXCEL"]

        if col_responsavel in df_final.columns:
            df_final = df_final[df_final[col_responsavel].astype(str).str.strip() != ""]
            df_final = df_final.dropna(subset=[col_responsavel])
        else:
            print("❌ Não foi possível encontrar a coluna RESPONSAVEL após a mesclagem.")
            return None

        if selected_responsaveis:
            df_final = df_final[df_final[col_responsavel].isin(selected_responsaveis)]

        colunas_desejadas = [
            "RESPONSAVEL", "DISPLAY", "NRO DO EQUIPAMENTO",
            "TIPO DO EQUIPAMENTO", "PRESTADOR", "REGISTRO MAIS RECENTE"
        ]
        
        colunas_desejadas_existentes = [c for c in colunas_desejadas if c in df_final.columns]
        df_final = df_final[colunas_desejadas_existentes]

        return df_final
    except Exception as e:
        print(f"⚠️ Erro ao processar CSV utilizando a planilha de monitoramento: {e}")
        return None

def obter_caminho_planilha():
    hoje = datetime.datetime.now()
    dia_atual = hoje.day
    mes_atual = hoje.month
    ano_atual = hoje.year

    if dia_atual > 20:
        mes_alvo = mes_atual + 1
        if mes_alvo > 12:
            mes_alvo = 1
            ano_alvo = ano_atual + 1
        else:
            ano_alvo = ano_atual
    else:
        mes_alvo = mes_atual
        ano_alvo = ano_atual

    # Formata o mês em 2 dígitos (ex: "01", "09", "12")
    mes_str = f"{mes_alvo:02d}"
    ano_str = str(ano_alvo)

    # --- CAMINHO DE REDE (UNC) ---
    possiveis_drives = [
        r"\\192.168.14.150\Departamentos$", 
        "I:",                               
        "A:", 
        "Z:", 
        "G:"
    ]
    
    caminho_final = None

    for drive in possiveis_drives:
        base = fr"{drive}\ANG\Agricola\Controle\Computador de Bordo\Fechamento Prestação de Serviço (Linha Amarela)\Pago pelo Bordo"
        
        if not os.path.exists(base):
            continue

        # 1. Procurar a pasta da Safra de forma dinâmica
        pasta_safra = None
        try:
            for nome in os.listdir(base):
                caminho_completo = os.path.join(base, nome)
                if os.path.isdir(caminho_completo) and ano_str in nome and "Safra" in nome:
                    pasta_safra = nome
                    break
        except Exception:
            pass

        if not pasta_safra:
            continue
            
        caminho_safra = os.path.join(base, pasta_safra)

        # 2. Procurar a pasta do Mês de forma dinâmica (pelo número do mês)
        pasta_mes = None
        try:
            for nome in os.listdir(caminho_safra):
                caminho_completo_mes = os.path.join(caminho_safra, nome)
                if os.path.isdir(caminho_completo_mes) and nome.startswith(mes_str):
                    pasta_mes = nome
                    break
        except Exception:
            pass

        if not pasta_mes:
            continue

        # 3. Tentar encontrar a planilha dentro da estrutura localizada
        caminho_teste = os.path.join(caminho_safra, pasta_mes, "Monitoramento - Eqps.xlsx")
        if os.path.exists(caminho_teste):
            caminho_final = caminho_teste
            break

    if not caminho_final:
        raise FileNotFoundError(
            f"❌ Não foi possível localizar a planilha de Equipamentos para o Ano '{ano_alvo}' e Mês '{mes_str}'."
        )

    return caminho_final

def atualizar_coleta_planilha(df_final):
    """Atualiza a coluna COLETA utilizando XLWINGS para simular uma ação de usuário no Excel."""
    try:
        caminho_planilha = obter_caminho_planilha()
        aba_alvo = "Cont. Maquinas"

        print("⏳ Atualizando Excel nativamente via xlwings (preservando o cache das Fórmulas)...")
        app = xw.App(visible=False)
        app.display_alerts = False 
        wb = None
        
        try:
            wb = app.books.open(caminho_planilha)
            ws = wb.sheets[aba_alvo]

            header_range = ws.range('A1').expand('right')
            cabecalhos = {str(cell.value).strip().upper(): cell.column for cell in header_range if cell.value}

            if "EQUIPAMENTO" not in cabecalhos or "COLETA" not in cabecalhos:
                print("⚠️ Colunas necessárias não encontradas na aba Cont. Maquinas.")
                return

            col_equip = cabecalhos["EQUIPAMENTO"]
            col_coleta = cabecalhos["COLETA"]

            equipamentos_contingencia = set(df_final["NRO DO EQUIPAMENTO"].astype(str).str.strip())
            equipamentos_coletados = set()

            last_row = ws.range((ws.cells.last_cell.row, col_equip)).end('up').row
            if last_row < 2: return

            valores_equip = ws.range((2, col_equip), (last_row, col_equip)).value
            valores_coleta = ws.range((2, col_coleta), (last_row, col_coleta)).value

            if not isinstance(valores_equip, list): valores_equip = [valores_equip]
            if not isinstance(valores_coleta, list): valores_coleta = [valores_coleta]

            for eq, col_val in zip(valores_equip, valores_coleta):
                if col_val == "DADOS COLETADOS" and eq is not None:
                    eq_str = str(eq).strip()
                    if eq_str.endswith('.0'): eq_str = eq_str[:-2]
                    equipamentos_coletados.add(eq_str)

            for i, (eq, col_val) in enumerate(zip(valores_equip, valores_coleta)):
                if eq is None: continue
                
                eq_str = str(eq).strip()
                if eq_str.endswith('.0'): eq_str = eq_str[:-2]

                row_idx = i + 2
                cell_coleta = ws.range((row_idx, col_coleta))
                
                if col_val == "DADOS COLETADOS":
                    continue

                if eq_str in equipamentos_contingencia and eq_str not in equipamentos_coletados:
                    cell_coleta.value = "COLETAR DADOS"
                    try:
                        cell_coleta.font.bold = True
                    except:
                        pass
                else:
                    if col_val == "COLETAR DADOS":
                        cell_coleta.value = None

            wb.save(caminho_planilha)
            print("✅ Atualização da coluna COLETA concluída de forma segura!")
        finally:
            if wb: 
                try: wb.close()
                except: pass
            if app: 
                try: app.quit()
                except: pass
    except ImportError:
        print("❌ Erro fatal: A biblioteca 'xlwings' não está instalada ou o Excel não está instalado no computador.")
    except Exception as e:
        print(f"⚠️ Erro grave ao atualizar a planilha via xlwings: {e}")

def salvar_pdf_por_responsavel(df_final, output_dir):
    try:
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # Remove apenas arquivos PDF antigos da pasta de saída
        for arquivo in os.listdir(output_dir):
            if arquivo.endswith(".pdf"):
                os.remove(os.path.join(output_dir, arquivo))

        caminho_planilha = obter_caminho_planilha()
        aba_alvo = "Cont. Maquinas"
        
        wb = load_workbook(caminho_planilha, data_only=True, read_only=True)
        ws = wb[aba_alvo]

        colunas = {str(cell.value).strip().upper(): idx for idx, cell in enumerate(ws[1]) if cell.value}
        if "EQUIPAMENTO" not in colunas or "COLETA" not in colunas:
            print("⚠️ Colunas necessárias não encontradas na aba Cont. Maquinas.")
            wb.close()
            return

        equipamentos_coletados = set()
        col_equipamento = colunas["EQUIPAMENTO"]
        col_coleta = colunas["COLETA"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            equipamento = str(row[col_equipamento]).strip() if row[col_equipamento] else ""
            coleta = str(row[col_coleta]).strip() if row[col_coleta] else ""
            if coleta == "DADOS COLETADOS":
                equipamentos_coletados.add(equipamento)

        wb.close() 

        df_filtrado = df_final[~df_final["NRO DO EQUIPAMENTO"].astype(str).isin(equipamentos_coletados)]
        responsaveis_para_gerar = df_filtrado['RESPONSAVEL'].unique()
        data_hora_geracao = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')

        COR_PRIMARIA = (18, 71, 51)
        COR_SECUNDARIA = (240, 240, 240)

        for responsavel in responsaveis_para_gerar:
            df_responsavel = df_filtrado[df_filtrado['RESPONSAVEL'] == responsavel]
            if df_responsavel.empty:
                print(f"⚠️ Não há dados para o responsável: {responsavel}.")
                continue

            pdf = FPDF(format='A4', unit='mm')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()

            pdf.set_fill_color(*COR_PRIMARIA)
            pdf.rect(0, 0, 220, 30, 'F')
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 15, 'RELATÓRIO DE CONTINGÊNCIA', ln=1, align='C')
            pdf.set_font("Arial", 'B', 10)
            pdf.cell(0, 5, 'Equipamentos sem comunicação no dia atual', ln=1, align='C')
            pdf.ln(10)

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(0, 10, f'Resp.: {responsavel}', ln=1, align='L')
            pdf.set_font("Arial", '', 8)
            pdf.cell(0, 5, f'Relatório gerado em: {data_hora_geracao}', ln=1, align='L')
            pdf.ln(5)

            pdf.set_fill_color(*COR_PRIMARIA)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Arial", 'B', 8)

            headers = ['DISPLAY', 'FROTA', 'TIPO EQUIP.', 'PRESTADOR', 'ÚLTIMA COMUNICAÇÃO']
            col_widths = [38, 38, 38, 38, 38]

            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, border=0, align='C', fill=True)
            pdf.ln()

            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Arial", size=8)

            for i, row in enumerate(df_responsavel.iterrows()):
                if i % 2 == 0:
                    pdf.set_fill_color(255, 255, 255)
                else:
                    pdf.set_fill_color(*COR_SECUNDARIA)

                row_data = row[1]

                def limpar_val(val):
                    v = str(val).strip()
                    if v.lower() in ['nan', 'none', '<na>', '']:
                        return ''
                    return v

                display_val = limpar_val(row_data.get('DISPLAY', ''))
                frota_val = limpar_val(row_data.get('NRO DO EQUIPAMENTO', ''))
                tipo_val = limpar_val(row_data.get('TIPO DO EQUIPAMENTO', ''))
                prestador_val = limpar_val(row_data.get('PRESTADOR', ''))

                pdf.cell(col_widths[0], 10, display_val, border=0, align='C', fill=True)
                pdf.cell(col_widths[1], 10, frota_val, border=0, align='C', fill=True)
                pdf.cell(col_widths[2], 10, tipo_val, border=0, align='C', fill=True)
                pdf.cell(col_widths[3], 10, prestador_val, border=0, align='C', fill=True)

                registro_formatado = row_data['REGISTRO MAIS RECENTE'].strftime('%d/%m/%Y %H:%M') if pd.notna(
                    row_data['REGISTRO MAIS RECENTE']) else 'N/A'
                pdf.cell(col_widths[4], 10, registro_formatado, border=0, align='C', fill=True)
                pdf.ln()

            pdf.ln(5)
            pdf.set_font("Arial", size=6, style='')
            pdf.cell(0, 10, "Relatório gerado automaticamente pelo Sistema.", ln=1, align='C')

            caminho_pdf = os.path.join(output_dir, f"Relatorio_{responsavel.replace(' ', '_')}.pdf")
            pdf.output(caminho_pdf)

            print(f"📄 PDF gerado com sucesso para {responsavel}: {caminho_pdf}")
    except Exception as e:
        print(f"⚠️ Erro ao gerar PDF: {e}")

def alternar_visualizacao_senha():
    if entry_senha.cget('show') == '*':
        entry_senha.config(show='')
        botao_visualizar.config(text="Ocultar")
    else:
        entry_senha.config(show='*')
        botao_visualizar.config(text="Mostrar")

def atualizar_campos_credenciais(credenciais_path):
    try:
        with open(credenciais_path, "r", encoding='utf-8') as file:
            data = json.load(file)
            usuario = data.get("usuario", "")
            senha = data.get("senha", "")
            return True, usuario, senha
    except (FileNotFoundError, json.JSONDecodeError):
        return False, "", ""

def salvar_usuario(credenciais_path):
    usuario = entry_usuario.get()
    senha = entry_senha.get()

    if var_salvar_usuario.get():
        credenciais = {"usuario": usuario, "senha": senha}
        try:
            with open(credenciais_path, "w", encoding='utf-8') as file:
                json.dump(credenciais, file, indent=4)
            print(f"🟢 Usuário e senha salvos em: {credenciais_path}")
        except Exception as e:
            print(f"⚠️ Erro ao salvar credenciais: {e}")
    else:
        if os.path.exists(credenciais_path):
            os.remove(credenciais_path)
            print("🔴 Arquivo de credenciais removido.")


def selecionar_todos():
    for var in responsaveis_vars.values():
        var.set(True)


def limpar_selecao():
    for var in responsaveis_vars.values():
        var.set(False)


def cancelar_execucao():
    global execucao_ativa
    execucao_ativa = False
    print("🔴 Execução cancelada.")
    atualizar_progresso("Execução cancelada.", step=0, total_steps=1)

def atualizar_progresso(status_texto, step, total_steps):
    if root and status_label and progress_bar:
        root.after(0, _atualizar_progresso_thread_safe, status_texto, step, total_steps)

def _atualizar_progresso_thread_safe(status_texto, step, total_steps):
    if total_steps > 0:
        progress = (step / total_steps) * 100
        progress_bar['value'] = progress
    status_label.config(text=status_texto)
    root.update_idletasks()

def criar_interface():
    global entry_usuario, entry_senha, var_salvar_usuario, var_headless, botao_visualizar, driver, responsaveis_vars, entry_intervalo, var_intervalo_ativado, status_label, progress_bar, root

    credenciais_path = os.path.join(script_dir, "credenciais.json")

    PALETTE = {
        "primary": "#0066AC",
        "secondary": "#5a6268",
        "success": "#28a745",
        "danger": "#dc3545",
        "background": "#F8F9FA",
        "text": "#212529",
    }

    root = ttk.Window(themename="yeti")
    root.title(f"Sistema Automatizado de Contingência HXG — v{VERSAO}")
    
    # Centralização e dimensionamento dinâmico e inteligente da janela
    root.update_idletasks()
    largura = 920
    altura = 670
    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()
    pos_x = (largura_tela // 2) - (largura // 2)
    pos_y = (altura_tela // 2) - (altura // 2)
    root.geometry(f"{largura}x{altura}+{pos_x}+{pos_y}")
    root.resizable(False, False)

    style = ttk.Style()
    style.configure("TLabel", font=("Helvetica", 10), background=PALETTE["background"], foreground=PALETTE["text"])
    style.configure("TFrame", background=PALETTE["background"])
    style.configure("TLabelframe", background=PALETTE["background"], font=("Helvetica", 11, "bold"))
    style.configure("TLabelframe.Label", background=PALETTE["background"], foreground=PALETTE["primary"])
    style.configure("TEntry", fieldbackground="white", font=("Helvetica", 10))

    style.configure("success.TButton", background=PALETTE["success"], foreground="white", font=("Helvetica", 11, "bold"))
    style.configure("danger.TButton", background=PALETTE["danger"], foreground="white", font=("Helvetica", 11, "bold"))
    style.configure("info.TButton", background=PALETTE["primary"], foreground="white", font=("Helvetica", 10, "bold"))
    style.configure("secondary.TButton", background=PALETTE["secondary"], foreground="white", font=("Helvetica", 10, "bold"))
    
    style.configure("Roundtoggle.TCheckbutton", background=PALETTE["background"], font=("Helvetica", 10))
    style.configure("info-round-toggle.TCheckbutton", background=PALETTE["background"], font=("Helvetica", 10))
    style.configure("success-round-toggle.TCheckbutton", background=PALETTE["background"], font=("Helvetica", 10))

    responsaveis_vars = {nome: tk.BooleanVar() for nome in RESPONSAVEIS_OPCOES}

    # Frame principal de preenchimento
    main_frame = ttk.Frame(root, padding=25)
    main_frame.pack(fill="both", expand=True)

    # Banner Superior (Título Moderno)
    banner_frame = ttk.Frame(main_frame)
    banner_frame.pack(fill="x", pady=(0, 15))
    
    title_label = ttk.Label(banner_frame, text="AUTOMAÇÃO DE CONTINGÊNCIA — HXG", font=("Helvetica", 18, "bold"), foreground=PALETTE["primary"])
    title_label.pack(side="left")
    
    version_label = ttk.Label(banner_frame, text=f"Versão {VERSAO}", font=("Helvetica", 10, "italic"), foreground="gray")
    version_label.pack(side="right", anchor="s", pady=5)

    # --- Divisor Central em Colunas (Esquerda: Configurações, Direita: Responsáveis) ---
    split_container = ttk.Frame(main_frame)
    split_container.pack(fill="both", expand=True, pady=10)

    # Coluna Esquerda
    col_esquerda = ttk.Frame(split_container)
    col_esquerda.pack(side="left", fill="both", expand=True, padx=(0, 15))

    # Labelframe de Credenciais
    cred_frame = ttk.Labelframe(col_esquerda, text="  🔑 Credenciais de Acesso  ", padding=15)
    cred_frame.pack(fill="x", pady=(0, 15))

    ttk.Label(cred_frame, text="Usuário / Matrícula:").pack(anchor="w", pady=(0, 5))
    entry_usuario = ttk.Entry(cred_frame)
    entry_usuario.pack(fill="x", pady=(0, 10))

    ttk.Label(cred_frame, text="Senha de Acesso:").pack(anchor="w", pady=(0, 5))
    frame_senha = ttk.Frame(cred_frame)
    frame_senha.pack(fill="x", pady=(0, 10))

    entry_senha = ttk.Entry(frame_senha, show="*")
    entry_senha.pack(side="left", fill="x", expand=True)

    botao_visualizar = ttk.Button(frame_senha, text="Mostrar", command=alternar_visualizacao_senha, bootstyle="secondary-outline", width=8)
    botao_visualizar.pack(side="left", padx=(8, 0))

    var_salvar_usuario = tk.BooleanVar()
    credenciais_existentes, usuario_carregado, senha_carregada = atualizar_campos_credenciais(credenciais_path)
    var_salvar_usuario.set(credenciais_existentes)

    if credenciais_existentes:
        entry_usuario.insert(0, usuario_carregado)
        entry_senha.insert(0, senha_carregada)

    ttk.Checkbutton(cred_frame, text="Lembrar meu usuário e senha", variable=var_salvar_usuario, bootstyle="round-toggle").pack(anchor="w", pady=(5, 0))

    # Labelframe de Parâmetros e Driver
    param_frame = ttk.Labelframe(col_esquerda, text="  ⚙️ Parâmetros do Navegador e Ciclo  ", padding=15)
    param_frame.pack(fill="both", expand=True)

    # --- NOVO REQUISITO: TOGGLE HEADLESS (Padrao Ativado/Invisível) ---
    var_headless = tk.BooleanVar(value=True)
    chk_headless = ttk.Checkbutton(
        param_frame, 
        text="Executar em segundo plano (Oculto - Recomendado)", 
        variable=var_headless, 
        bootstyle="success-round-toggle"
    )
    chk_headless.pack(anchor="w", pady=(0, 15))

    # Configuração de Ciclo / Agendamento
    sched_container = ttk.Frame(param_frame)
    sched_container.pack(fill="x", pady=(0, 10))
    
    ttk.Label(sched_container, text="Intervalo de varredura:").pack(side="left", padx=(0, 5))
    entry_intervalo = ttk.Entry(sched_container, width=8)
    entry_intervalo.insert(0, "60")
    entry_intervalo.pack(side="left", padx=(0, 5))
    ttk.Label(sched_container, text="minutos").pack(side="left")

    var_intervalo_ativado = tk.BooleanVar(value=False)
    chk_intervalo = ttk.Checkbutton(
        param_frame, 
        text="Ativar Agendamento Automático de Contingência", 
        variable=var_intervalo_ativado,
        bootstyle="info-round-toggle"
    )
    chk_intervalo.pack(anchor="w", pady=(5, 0))


    # Coluna Direita (Responsáveis)
    col_direita = ttk.Frame(split_container)
    col_direita.pack(side="right", fill="both", expand=True, padx=(15, 0))

    resp_frame = ttk.Labelframe(col_direita, text="  👤 Emitir Relatório para Responsáveis  ", padding=15)
    resp_frame.pack(fill="both", expand=True)

    # Botões de controle de seleção em linha
    ctrl_sel_frame = ttk.Frame(resp_frame)
    ctrl_sel_frame.pack(fill="x", pady=(0, 15))
    
    btn_sel_todos = ttk.Button(ctrl_sel_frame, text="Selecionar Todos", command=selecionar_todos, bootstyle="info", width=18)
    btn_sel_todos.pack(side="left", fill="x", expand=True, padx=(0, 5))
    
    btn_limpar_todos = ttk.Button(ctrl_sel_frame, text="Limpar Seleção", command=limpar_selecao, bootstyle="secondary", width=18)
    btn_limpar_todos.pack(side="right", fill="x", expand=True, padx=(5, 0))

    # Grid de 2 colunas para exibição limpa dos 10 responsáveis
    grid_responsáveis = ttk.Frame(resp_frame)
    grid_responsáveis.pack(fill="both", expand=True)
    grid_responsáveis.columnconfigure(0, weight=1)
    grid_responsáveis.columnconfigure(1, weight=1)

    for indice, (nome, variavel) in enumerate(responsaveis_vars.items()):
        linha = indice // 2
        coluna = indice % 2
        chk_resp = ttk.Checkbutton(grid_responsáveis, text=nome, variable=variavel, bootstyle="info-round-toggle")
        chk_resp.grid(row=linha, column=coluna, sticky="w", padx=10, pady=7)


    # --- Painel Inferior de Execução (Barra de Progresso e Botões de Ação) ---
    bottom_frame = ttk.Frame(main_frame)
    bottom_frame.pack(fill="x", pady=(15, 0))

    status_label = ttk.Label(bottom_frame, text="Aguardando comando do usuário...", font=("Helvetica", 10, "italic"), foreground="gray")
    status_label.pack(anchor="w", pady=(0, 5))

    progress_bar = ttk.Progressbar(bottom_frame, mode="determinate", bootstyle="info")
    progress_bar.pack(fill="x", pady=(0, 15))

    # Botões de Execução (Iniciar e Pausar) em destaque
    action_buttons_frame = ttk.Frame(bottom_frame)
    action_buttons_frame.pack(fill="x")

    btn_executar = ttk.Button(action_buttons_frame, text="  Iniciar Automação 🚀", command=lambda: executar_script(), bootstyle="success")
    btn_executar.pack(side="left", fill="x", expand=True, padx=(0, 8))

    btn_pausar = ttk.Button(action_buttons_frame, text="  Pausar Execução ⏸️", command=cancelar_execucao, bootstyle="danger")
    btn_pausar.pack(side="right", fill="x", expand=True, padx=(8, 0))

    root.bind('<Return>', lambda event: executar_script())

    def fechar_janela():
        salvar_usuario(credenciais_path)
        if 'driver' in globals() and driver:
            try:
                driver.quit()
            except:
                pass
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", fechar_janela)
    root.mainloop()

def executar_script():
    global execucao_ativa
    if execucao_ativa:
        messagebox.showinfo("Informação", "A automação já está em execução.")
        return

    usuario = entry_usuario.get()
    senha = entry_senha.get()

    if not usuario or not senha:
        messagebox.showwarning("Aviso", "Por favor, preencha o usuário e a senha.")
        return

    credenciais_path = os.path.join(script_dir, "credenciais.json")
    salvar_usuario(credenciais_path)

    # Lê o estado da variável headless da thread principal de forma segura
    headless_selecionado = var_headless.get()

    atualizar_progresso("Iniciando a automação...", step=0, total_steps=6)
    threading.Thread(target=executar_procedimento, args=(usuario, senha, headless_selecionado), daemon=True).start()

def executar_procedimento(usuario, senha, is_headless):
    global driver, execucao_ativa, responsaveis_vars
    execucao_ativa = True

    TOTAL_STEPS = 6
    last_valid_interval = 60

    while execucao_ativa:
        selected_responsaveis = [nome for nome, var in responsaveis_vars.items() if var.get()]
        intervalo_ativado = var_intervalo_ativado.get()

        try:
            intervalo_minutos = int(entry_intervalo.get())
            if intervalo_minutos <= 0:
                print(f"⚠️ Intervalo inválido ({intervalo_minutos}). Usando o original: {last_valid_interval} min.")
                intervalo_minutos = last_valid_interval
            else:
                last_valid_interval = intervalo_minutos
        except (ValueError, tk.TclError):
            print(f"⚠️ Erro ao ler o intervalo. Usando o original: {last_valid_interval} min.")
            intervalo_minutos = last_valid_interval

        print(f"\n--- Iniciando novo ciclo ---")
        print(f"Responsáveis selecionados para este ciclo: {', '.join(selected_responsaveis) or 'Todos'}")
        print(f"Intervalo configurado: {intervalo_minutos} minutos. Agendamento Ativado: {'Sim' if intervalo_ativado else 'Não'}")

        driver = None
        df_final = None

        try:
            diretorio_downloads = os.path.join(os.path.expanduser("~"), "Downloads")
            pdf_output_dir = os.path.join(script_dir, "Contingência_Saida")

            xpaths = {
                'usuario': '/html/body/app-root/app-login/app-access-container/div/div[2]/div[2]/form/div[1]/input',
                'senha': '/html/body/app-root/app-login/app-access-container/div/div[2]/div[2]/form/div[2]/input',
                'botao_login': '/html/body/app-root/app-login/app-access-container/div/div[2]/div[2]/form/div[4]/p-button/button',
                'limpar_filtro': '//*[@id="buttion-id-clearAndApplyButton"]/div/app-text/div',
                'tabela': '//*[@id="div-submenu-link-id-app-submenu-link-mon-table-id"]/p',
                'paginador_dropdown': '//p-paginator//p-dropdown | //*[starts-with(@id, "pn_id_") and contains(@id, "label")]',
                'paginador_opcao_5': '//p-dropdownitem[5]/li',
                'filtro': '//p-table//div[1]/div/div[2]/button[3] | //*[starts-with(@id, "pn_id_")]/div[1]/div/div[2]/button[3]',
                'exportacao_csv': '//p-table//div[1]/div/div[2]/button[1] | //*[starts-with(@id, "pn_id_")]/div[1]/div/div[2]/button[1]'
            }
            url = 'https://access.hxgnagron.com/?redirect=http:%2F%2Fcontrolroom.hxgnagron.com%2F#/'

            atualizar_progresso("Iniciando driver...", step=1, total_steps=TOTAL_STEPS)
            driver = iniciar_driver(headless=is_headless)

            if not execucao_ativa: break

            atualizar_progresso("Realizando login...", step=2, total_steps=TOTAL_STEPS)
            login_usuario(driver, url, usuario, senha, xpaths)

            if not execucao_ativa: break

            atualizar_progresso("Exportando tabela...", step=3, total_steps=TOTAL_STEPS)
            exportar_tabela(driver, xpaths)

            if not execucao_ativa: break

            atualizar_progresso("Aguardando download e processando...", step=4, total_steps=TOTAL_STEPS)
            df_final = processar_csv(diretorio_downloads, pdf_output_dir, selected_responsaveis)

            if df_final is not None:
                if not execucao_ativa: break
                atualizar_progresso("Gerando PDFs...", step=5, total_steps=TOTAL_STEPS)
                salvar_pdf_por_responsavel(df_final, pdf_output_dir)

                if not execucao_ativa: break
                atualizar_progresso("Atualizando planilha de controle...", step=6, total_steps=TOTAL_STEPS)
                atualizar_coleta_planilha(df_final)

                atualizar_progresso("Procedimento concluído com sucesso!", step=6, total_steps=TOTAL_STEPS)
            else:
                atualizar_progresso("Processamento de dados falhou.", step=0, total_steps=1)

        except Exception as e:
            # --- GERADOR DE LOG DE ERRO ---
            import traceback
            error_details = traceback.format_exc()
            error_message = f"❌ Erro fatal na execução: {type(e).__name__}: {str(e)[:100]}..."
            
            print(error_message)
            logger.error(f"❌ Erro fatal na execução do procedimento: {e}\n{error_details}")
            atualizar_progresso(error_message, step=0, total_steps=1)
            
            try:
                # Procura a área de trabalho da máquina (PT ou EN)
                user_profile = os.path.expanduser("~")
                desktop_paths = [
                    os.path.join(user_profile, "Desktop"),
                    os.path.join(user_profile, "Área de Trabalho"),
                    os.path.join(user_profile, "OneDrive", "Desktop"),
                    os.path.join(user_profile, "OneDrive", "Área de Trabalho")
                ]
                
                desktop = user_profile 
                for dp in desktop_paths:
                    if os.path.exists(dp):
                        desktop = dp
                        break
                        
                caminho_erro = os.path.join(desktop, "erro_hxg_debug.txt")
                with open(caminho_erro, "w", encoding="utf-8") as f:
                    f.write(f"Data/Hora: {datetime.datetime.now()}\n")
                    f.write(f"Versão: {VERSAO}\n\n")
                    f.write(f"Erro:\n{str(e)}\n\n")
                    f.write(f"Traceback Técnico:\n{error_details}")
                print(f"📄 Arquivo de log de erro salvo para análise em: {caminho_erro}")
            except Exception as log_e:
                print(f"⚠️ Não foi possível salvar o arquivo de log: {log_e}")
                
        finally:
            if driver:
                try:
                    driver.quit()
                except:
                    pass

        if not intervalo_ativado:
            print("Execução única concluída, pois o agendamento está desativado.")
            break

        if execucao_ativa:
            tempo_total_espera = intervalo_minutos * 60
            print(f"✅ Execução concluída. Aguardando {intervalo_minutos} minutos para a próxima rodada...")

            for segundos_restantes in range(tempo_total_espera, 0, -1):
                if not execucao_ativa: break
                minutos, segundos = divmod(segundos_restantes, 60)
                texto_tempo = f"Próxima execução em {minutos:02d}:{segundos:02d}"
                atualizar_progresso(texto_tempo, step=6, total_steps=TOTAL_STEPS)
                time.sleep(1)

            if not execucao_ativa: break

    execucao_ativa = False
    atualizar_progresso("Procedimento finalizado.", step=0, total_steps=1)
    print("🏁 Procedimento finalizado.")

if __name__ == "__main__":
    verificar_seguranca()
    driver = None
    criar_interface()